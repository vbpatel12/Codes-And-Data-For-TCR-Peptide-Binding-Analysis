import os
import shutil
import openpyxl as xl
import random
import rdkit
from rdkit import Chem
import sys
import os
from rdkit.Chem import rdmolfiles
from rdkit.Chem.rdmolfiles import MolFromPDBFile
from rdkit.Chem.rdmolfiles import MolToPDBFile
from rdkit.Chem import rdMolTransforms
from rdkit.Chem.rdMolTransforms import GetDihedralDeg, GetAngleDeg
from rdkit.Chem.rdMolTransforms import SetDihedralDeg, SetAngleDeg
from rdkit.Chem.rdMolTransforms import SetBondLength
from rdkit.Chem.AllChem import EmbedMultipleConfs, MMFFOptimizeMoleculeConfs

#make the solvare_peptide_files and define the path to the main folder
path = r'C:\People\Vini2\AddProt_exp\add_input_files'
directory_solvate = 'Solvate_peptide_files2'
parent_dir2 = r'C:\People\Vini2'
path4 = os.path.join(parent_dir2, directory_solvate)
os.mkdir(path4)

print('solvate_peptide_files directory made')
#Copy the peptide files and Addprot.exe to the main folder
files = os.listdir('C:\People\Vini2\AddProt_exp\Peptide')
for file in files:
    original_src = ('C:\People\Vini2\AddProt_exp\Peptide\\') + file
    shutil.copy(original_src, path)
original_src_addprot = 'C:\People\Vini2\AddProt_exp\exe\AddProt_2May.exe'
shutil.copy(original_src_addprot, path)
print('peptide files and Addprot.exe copied to add_input_files')
# open the excel file:

wb = xl.load_workbook('C:\People\Vini2\AddProt_exp\AddProt_Master.xlsx', data_only=True)
sheet = wb['Sheet5']
# using for loop- assign variable to each cells of different columns
for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 2)
    cell2 = sheet.cell(row, 3)
    cell3 = sheet.cell(row, 8)
    cell4 = sheet.cell(row, 7)
    cell5 = sheet.cell(row, 9)
    cell6 = sheet.cell(row, 10)
    cell7 = sheet.cell(row, 11)
    cell8 = sheet.cell(row, 12)
    cell9 = sheet.cell(row, 13)
    cell10 = sheet.cell(row, 14)
    cell11 = sheet.cell(row, 15)
    cell12 = sheet.cell(row, 16)
    cell13 = sheet.cell(row, 5)
    cell14 = sheet.cell(row,17)
    cell15 = sheet.cell(row,18)
    # retrieve the values from each cells and assigning variables to the cells
    sequence = cell2.value
    pdb_names = cell.value
    peptide_varied = str(cell4.value)
    fixed_pep_1 = cell3.value
    fixed_pep2 = cell5.value
    angle1 = str(cell6.value)
    angle2 = str(cell7.value)
    increment = str(cell8.value)
    error1 = str(cell9.value)
    error2 = str(cell10.value)
    error3 = str(cell11.value)
    clashdist = str(cell12.value)
    file_names = str(pdb_names)
    peptide_len = cell13.value
    vary_pep_num1 = str(cell14.value)
    vary_pep_num2 = str(cell15.value)

    # Create the directories with the pdb_ids as the directory name
    directory = file_names
    parent_dir = 'C:\People\Vini2\AddProt_exp'
    path2 = os.path.join(parent_dir, directory)
    #Create Vary folder within each directory as run.bat recognizes this file path
    directory2 = 'Vary'
    path3 = os.path.join(path2, directory2)
    os.mkdir(path2)
    os.mkdir(path3)
    # Create the directories with pdb names in the solvate peptide files folder
    directory_inside_solvate_peptide_files = file_names
    path5 = os.path.join(path4, directory_inside_solvate_peptide_files)
    os.mkdir(path5)

    # format for in_peptide.txt file:
    line1 = ' ' + ' ' + '0' + ' ' + ' ' + '2' + ' ' + ' ' + '0' + ' ' + ' ' + '1' + ' ' + ' ' + '0' + ' ' + ' ' + '0' + '\n'
    line2 = file_names + '_peptide.pdb' + '\n'

    for i in range(len(sequence)):
        variedseq = sequence[fixed_pep_1: fixed_pep2 - 1]
        a = len(variedseq)
        newseq = sequence[0:fixed_pep_1] + a * '-' + sequence[fixed_pep2 - 1:peptide_len]
        line3 = newseq + '\n'
    fixed_pep_1 = str(cell3.value)
    fixed_pep2 = str(cell5.value)
    line4 = '    ' + fixed_pep_1 + '    ' + peptide_varied + '    ' + '0.0 0' + '    ' + fixed_pep2 + '\n'
    line5 = variedseq + '\n'
    line6 = '    ' + angle1 + '  ' + angle2 + '   ' + increment + '\n'
    line7 = '    ' + error1 + '   ' + error2 + '.0' + '   ' + error3 + '.0'+'    '+ clashdist
    data = line1 + line2 + line3 + line4 + line5 + line6 * 2* int(peptide_varied)+ line7

    # create in_peptide.txt file in each directory:

    output_file = file_names + '_in_peptide.txt'
    new_path = os.path.join(path, output_file)
    with open(new_path, 'w') as f:
        for line in data:
            f.write(line)
    # Run the Addprot.exe and generate the conformations
    # The conformations will be moved to respective folders along with data.out.txt file
    files = os.listdir(r'C:\People\Vini2\AddProt_exp\add_input_files')
    for file in files:
        if file == output_file:
            old_name = new_path
            new_name = r'C:\People\Vini2\AddProt_exp\add_input_files\in.txt'
            os.rename(old_name, new_name)
            os.system(r'C:\People\Vini2\AddProt_exp\add_input_files\AddProt_2May.exe')
            os.rename(new_name, old_name)
            shutil.move(new_path, path3)
            old_name_pep_file = r'C:\People\Vini2\AddProt_exp\add_input_files\out_ap.pdb'
            new_name_pep = file_names + '_vary_peptide.pdb'
            new_name_pep_file = os.path.join(path, new_name_pep)
            os.rename(old_name_pep_file,new_name_pep_file)
            src_path = new_name_pep_file
            src_path2 = r'C:\People\Vini2\AddProt_exp\add_input_files\data_ap.out'
            shutil.copy(src_path, path3)
            shutil.move(src_path2, path3)
    print('conformation file made and moved to ' +file_names + ' ' + 'directory')

    #Edit run.bat file and move to respective folder

    input_file = r"C:\People\Vini2\AddProt_exp\run.bat"
    with open(input_file, 'r') as txt_addprot_file:
        lines = txt_addprot_file.readlines()
    for i, line in enumerate(lines):
        if line[18:22] == '1ao7':
            a = file_names
            lines[i] = lines[i].replace(lines[i][18:22], a)
    output_file2 = 'run.bat'
    run_bat_file = os.path.join(path3, output_file2)
    with open(run_bat_file, 'w') as f:
        for line in lines:
            f.write(line)

    # Copy the vary file(peptide conformation file) to Solvate_peptide_files
    vary_pep_file = file_names + '_vary_peptide.pdb'
    vary_pep_file_src = os.path.join(path3, vary_pep_file)
    shutil.copy(vary_pep_file_src, path5)
    path6 = os.path.join(path5, vary_pep_file)

    def get_num(pdb_name):
        input_file = pdb_name + '_vary_peptide.pdb'
        # read as txt file
        with open(input_file, 'r') as pdb_file:
            lines = pdb_file.readlines()

        N = 0
        for line in lines:
            if line[0:3] == 'TER':
                N += 1
        return N


    def change_number(fix_1, vary, fix_2):
        """
        Change the num of the vary part.
        """
        start = int(fix_1[-1][24:26]) + 1
        end = int(fix_2[0][24:26]) - 1
        last = vary[0][24:26]
        now = str(start)
        for i in range(len(vary)):
            line = vary[i]
            if line[24:26] == last:
                line = line[:24] + " " + now + line[26:]
            else:
                now = str(int(now) + 1)
                line = line[:24] + " " + now + line[26:]
                last = str(int(last) + 1)
            vary[i] = line
        return vary

    def run(pdb_name, rand):
        input_file = pdb_name + '_vary_peptide.pdb'
        # read as txt file
        with open(input_file, 'r') as pdb_file:
            lines = pdb_file.readlines()

        # format used to find the vary part
        # example: rand = 35
        # formatt_1 = 'TER   35'
        # formatt_2 = 'TER   36'
        format_1 = 'TER' + ' ' * (5 - len(str(rand))) + str(rand)
        format_2 = 'TER' + ' ' * (5 - len(str(rand + 1))) + str(rand + 1)

        # loop the lines -> find the format -> save the related index
        fix_vary_idx = []

        for i, line in enumerate(lines):
            if line[0:8] == format_1:
                vary_begin_idx = i + 1

            # ending index of vary part
            if line[0:8] == format_2:
                vary_end_idx = i

            # fix part
            if line.strip() == 'TER':
                fix_begin_idx = i + 1

            if line[0:3] != 'TER' and line[21].isalpha() == False:
                fix_vary_idx.append(i)

        # fix beginning + vary + fix ending
        fix_1 = lines[fix_begin_idx:fix_vary_idx[0]]
        vary = lines[vary_begin_idx:vary_end_idx]
        fix_2 = lines[fix_vary_idx[-1] + 1:]

        # Added step to change num
        vary = change_number(fix_1, vary, fix_2)

        result =fix_1 + vary + fix_2

        idx = [i + 1 for i in range(len(result))]

        for i in range(len(result)):
            result[i] = result[i][:-1] + '  1.00  0.00' + ' ' * 14 + '\n'
            result[i] = result[i][:8] + ' ' * (3 - len(str(idx[i]))) + str(idx[i]) + result[i][11:]
            result[i] = result[i].replace(result[i][21], "C")

            if result[i][12] != ' ':
                result[i] = result[i][:12] + ' ' + result[i][13:]

        result += ['TER\n']
        return result

    pdb_name = file_names

    N = get_num(pdb_name)
    # N = 104
    rand_list = random.sample(range(1, N-1), 20)
    print(N)
    print(rand_list)
    line_1 = str(N-1) + '\n'
    line_2 = str(rand_list)
    data_in_txt = line_1 + line_2
    in_txt_file = pdb_name + 'in.txt'
    with open(in_txt_file, 'w') as f:
        for line in data_in_txt:
            f.write(line)

    res_all = []
    for rand in rand_list:
        res = run(pdb_name, rand)
        res_all += res

    output_file = pdb_name + '_P_modified_test20.pdb'
    with open(output_file, 'w') as f:
        for line in res_all:
            f.write(line)

    input_file = pdb_name + '_P_modified_test20.pdb'
    with open(input_file, 'r') as f:
        lines = f.readlines()
    lines.insert(0, 'MODEL        1\n')
    #lines.replace([-1], 'MODEL        1\n')
    n= 1


    for i,line in enumerate(lines):

        if lines[i][0:3] == 'TER':
            n += 1

            lines[-1] = lines[-1].replace(lines[-1][0:3], '   ')
            lines[i] = lines[i].replace(lines[i][0:14], 'MODEL'+ ' '*(9-len(str(n))) + str(n)+ '\n')

        # if lines[i+1][0:5] == 'MODEL':
         #   lines[i] = lines[i].insert(lines[i][0:6],'ENDMDL')

       # lines.insert(i, "ENDMDL\n")
    lines = [line for line in lines if line.strip()]
    # lines.replace(-1,' \n')
        # open a new file to write the output
    with open("in_orig.pdb", "w") as f:
            # initialize a counter to keep track of the current model number
        model_num = 1
            # loop over the lines in the input file
        for i in range(len(lines)):
                # check if the line starts with "MODEL"
            if lines[i].startswith("MODEL"):
                    # if the model number is not 1, write "END" before the new model line
                if model_num != 1:
                    f.write("ENDMDL\n")
                    # write the model line and increment the model number
                f.write(lines[i])
                model_num += 1
            else:
                    # write the line
                f.write(lines[i])
            # write "END" after the last model
        f.write("ENDMDL\n")



    # Fixed start
    iatom = []  # H atom
    jatom = []  # N atom
    katom = []  # C atom
    latom = []  # O atom
    patom = []  # O atom - 6th Res
    qatom = []  # C atom - 6th Res
    ratom = []  # N atom - 7th Res
    satom = []
    Atindex_start = []
    # input_file = pdb_name + '_peptide_10conformers2.pdb'
    input_file = 'in_orig.pdb'
    with open(input_file, 'r') as pdb_file:
        lines = pdb_file.readlines()
    res = "MODEL" + "\n"
    res2 = "ENDMDL" + "\n"
    for i, line in enumerate(lines):
        if line[13:14] == 'O' and line[25:26] == fixed_pep_1:

            O = int(line[8:11]) - 1
            line = line.replace('O ', 'HX')
            O1 = line[0:]
            # print(O)
            latom.append(O)
            #outfile.write(line.replace("O   PHE C   3", "H   PHE C   3"))

            Atindex_start.append(O1)
        if line[13:15] == 'C ' and line[25:26] == fixed_pep_1:
            C = int(line[8:11]) - 1
            C1 = line[0:]
            katom.append(C)
            Atindex_start.append(res)
            Atindex_start.append(C1)
        if line[13:15] == 'N ' and line[25:26] == vary_pep_num1:
            N = int(line[8:11]) - 1
            N1 = line[0:]
            jatom.append(N)
            Atindex_start.append(N1)
        if line[13:15] == 'H ' and line[25:26] == vary_pep_num1:
            H = int(line[8:11]) - 1
            H1 = line[0:]
            iatom.append(H)
            Atindex_start.append(H1)

        if line[13:14] == 'O' and line[25:26] == vary_pep_num2:
            O = int(line[8:11]) - 1
            O2 = line[0:]
            # print(O)
            patom.append(O)

            Atindex_start.append(O2)
        if line[13:15] == 'C ' and line[25:26] == vary_pep_num2:
            C = int(line[8:11]) - 1
            C2 = line[0:]
            qatom.append(C)
            Atindex_start.append(C2)

        if line[13:15] == 'N ' and line[25:26] == fixed_pep2:
            N = int(line[8:11]) - 1
            N2 = line[0:]
            ratom.append(N)
            # print(N2)
            Atindex_start.append(N2)

        if lines[i][13:14] != 'H' and line[25:26] == fixed_pep2:
            if lines[i + 1][13:14] == 'H' and line[25:26] == fixed_pep2:
                H = int(line[8:11]) - 1
                H2 = lines[i + 1][0:]
                satom.append(H)

                Atindex_start.append(H2)

                Atindex_start.append(res2)

                print(Atindex_start)

            # print(latom, katom, jatom, iatom)
            # print(O1)
            # print(C1)
            # print(N1)
            # print(H1)
            print(Atindex_start)
            # res= "MODEL" + "\n" + Atindex_start

            # output_file = pdb_name + '_torsion.pdb'
        output_file = 'in_torsion.pdb'
        with open(output_file, 'w') as f:
            for line in Atindex_start:
                f.write(line)

    # mol = MolFromPDBFile(pdb_name + '_torsion.pdb',sanitize=False, removeHs=False)
    mol = MolFromPDBFile('in_torsion.pdb', sanitize=False, removeHs=False)
    print(mol)

    mol = MolFromPDBFile('in_torsion.pdb', sanitize=False, removeHs=False)
    chain = Chem.MolToPDBBlock(mol)
    print(chain)

    conf = mol.GetNumConformers()
    print(conf)

    for c in mol.GetConformers():
        atoms = [a for a in mol.GetAtoms()]
        for a in atoms:
            print(a.GetIdx(), a.GetSymbol())

    for c in mol.GetConformers():
        GetDihedralDeg(c, 3, 2, 0, 1)
        print(GetDihedralDeg(c, 3, 2, 0, 1))
    for c in mol.GetConformers():
        GetDihedralDeg(c, 5, 4, 6, 7)
        print(GetDihedralDeg(c, 5, 4, 6, 7))

    Chem.SanitizeMol(mol, Chem.SANITIZE_SYMMRINGS | Chem.SANITIZE_SETCONJUGATION | Chem.SANITIZE_SETHYBRIDIZATION)

    for c in mol.GetConformers():
        SetDihedralDeg(c, 3, 2, 0, 1, 180.0)
        print(SetDihedralDeg(c, 3, 2, 0, 1, 180.0))

    for c in mol.GetConformers():
        SetDihedralDeg(c, 5, 4, 6, 7, 180.0)
        print(SetDihedralDeg(c, 5, 4, 6, 7, 180.0))

    for c in mol.GetConformers():
        GetDihedralDeg(c, 3, 2, 0, 1)
        print(GetDihedralDeg(c, 3, 2, 0, 1))

    for c in mol.GetConformers():
        GetDihedralDeg(c, 5, 4, 6, 7)
        print(GetDihedralDeg(c, 5, 4, 6, 7))

    for c in mol.GetConformers():
        GetAngleDeg(c, 2, 0, 1)
        print(GetAngleDeg(c, 2, 0, 1))

    Chem.SanitizeMol(mol,Chem.SANITIZE_SYMMRINGS | Chem.SANITIZE_SETCONJUGATION | Chem.SANITIZE_SETHYBRIDIZATION)

    for c in mol.GetConformers():
        # SetBondLength(c,0,1,1.3)
        SetAngleDeg(c, 2, 0, 1, 120.0)
        print(SetAngleDeg(c, 2, 0, 1, 120.0))

    for c in mol.GetConformers():
        GetAngleDeg(c, 2, 0, 1)
        print(GetAngleDeg(c, 2, 0, 1))

    for c in mol.GetConformers():
        GetAngleDeg(c, 4, 6, 7)
        print(GetAngleDeg(c, 4, 6, 7))

    Chem.SanitizeMol(mol,Chem.SANITIZE_SYMMRINGS | Chem.SANITIZE_SETCONJUGATION | Chem.SANITIZE_SETHYBRIDIZATION)

    for c in mol.GetConformers():
        SetAngleDeg(c, 4, 6, 7, 120.0)
        print(SetAngleDeg(c, 4, 6, 7, 120.0))

    for c in mol.GetConformers():
        GetAngleDeg(c, 4, 6, 7)
        print(GetAngleDeg(c, 4, 6, 7))

    for c in mol.GetConformers():
        GetDihedralDeg(c, 3, 2, 0, 1)
        print(GetDihedralDeg(c, 3, 2, 0, 1))

    for c in mol.GetConformers():
        SetDihedralDeg(c, 5, 4, 6, 7, 180.0)
        print(SetDihedralDeg(c, 5, 4, 6, 7, 180.0))

    # MolToPDBFile(mol,"1AO7_pep_t6.pdb")
    MolToPDBFile(mol, "in_mod2.pdb")

    input_file = 'in_mod2.pdb'
    with open(input_file, 'r') as pdb_file:
        lines = pdb_file.readlines()

    #for i, line in enumerate(lines):
     #   if line[13:15] == 'HX' and line[25:26] == '3':
      #      print(line[13:15])
       #     line = line.replace('HX','O ')
        #    print(line[13:15])
    new_lines = []
    for line in lines:
        if line[13:15] == "HX":
            line = line.replace("HX", "O ")
        new_lines.append(line)

    with open('in_mod.pdb', 'w') as f:
        f.write(''.join(new_lines))
    # output_file = 'in_mod.pdb'
    #with open(output_file, 'w') as f:
     #   f.writelines(line)



    import subprocess

    subprocess.run(r'C:\People\Vini2\AddProt_exp\add_input_files\Torsion CalcE.exe')

    input_file = "out.pdb"

    with open(input_file, 'r') as pdb_file:
        lines = pdb_file.readlines()
        # print(lines)

    for i, line in enumerate(lines):

        if line[0:5] == 'MODEL':
            lines[i] = lines[i].replace(lines[i][0:14], 'TER'+' '*11)
            lines[0] = lines[0].replace(lines[0][0:5],'     ')
        if line[0:6] == 'ENDMDL':
            lines[i] = lines[i].replace(lines[i][0:6], '      ')
        # if line[77:78] != ' ':
         #   lines[i] = lines[i].replace(lines[i][77:78], " ")
    lines = [line for line in lines if line.strip()]
    lines = [line for line in lines if line.strip()]
    lines = [line for line in lines if line.strip()]

    output_file = pdb_name + 'out.pdb'
    with open(output_file, 'w') as f:
        for line in lines:
            f.write(line)
    os.remove(r'C:\People\Vini2\AddProt_exp\add_input_files\out.pdb')
    # os.remove(r'C:\People\Vini2\AddProt_exp\add_input_files\in_mod.pdb')
    # os.remove(r'C:\People\Vini2\AddProt_exp\add_input_files\in_orig.pdb')
    # os.remove(r'C:\People\Vini2\AddProt_exp\add_input_files\in_torsion.pdb')

    input_file = pdb_name + 'out.pdb'
    with open(input_file, 'r') as f:
        lines= f.readlines()

    for i, line in enumerate(lines):
        if line[0:4] == 'ATOM':
            lines[i] = lines[i].replace(lines[i][17:26], 'LIG L   1')

    output_file = pdb_name + '_p_20vary.pdb'
    with open(output_file, 'w') as f:
        for line in lines:
            f.write(line)

    input_file = pdb_name + '_p_20vary.pdb'
    with open(input_file,'r') as f:
        lines= f.readlines()

    output_file = pdb_name + '_pep_20vary.pdb'
    with open(output_file, 'w') as f:
        for line in lines:
            f.write(line)





    print(file_names + 'solvate_pep_file_made')
    path_output_vary_file = os.path.join(path, output_file)
    shutil.move(path_output_vary_file, path5)
    path_in_txt_file = os.path.join(path, in_txt_file)
    shutil.move(path_in_txt_file, path5)
    src_pause_bat = r'C:\People\Vini2\AddProt_exp\pause.bat'
    shutil.copy(src_pause_bat, path5)

















